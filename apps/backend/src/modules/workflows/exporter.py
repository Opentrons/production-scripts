from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules.workflows.models import WorkflowBomDifference, WorkflowBomIgnoredItem, WorkflowRun


STATUS_LABELS = {
    "missing_in_duro": "Duro 缺失",
    "extra_in_duro": "Duro 冗余",
    "quantity_mismatch": "数量差异",
    "quantity_unknown": "数量未知",
    "parent_bom_ignored": "父菜单未核对",
}

IGNORE_TYPE_LABELS = {
    "sop_product_keyword": "SOP 产品关键字",
    "part_number": "料号",
    "part_number_cleanup": "默认料号清洗",
    "parent_bom": "父菜单 BOM",
    "supply": "辅料",
}

BASE_HEADERS = [
    "差异类型",
    "料号",
    "物料名称",
    "SOP 数量",
    "Duro 数量",
    "差值（Duro - SOP）",
    "SOP 正文出现次数",
    "SOP 位置",
    "差异汇总说明",
    "累加过程",
    "Duro 下级 BOM",
    "Duro 路径",
]

IGNORED_HEADERS = [
    *BASE_HEADERS,
    "忽略类型",
    "忽略值",
    "忽略原因",
    "忽略时间",
    "规范后料号",
]


def build_workflow_run_workbook(run: WorkflowRun) -> tuple[bytes, str]:
    workbook = Workbook()
    difference_sheet = workbook.active
    difference_sheet.title = "差异明细"
    ignored_sheet = workbook.create_sheet("已忽略")

    report = run.report
    _write_sheet(
        difference_sheet,
        BASE_HEADERS,
        [_difference_row(item) for item in (report.differences if report else [])],
    )
    _write_sheet(
        ignored_sheet,
        IGNORED_HEADERS,
        [_ignored_row(item) for item in (report.ignored_items if report else [])],
    )

    workbook.properties.title = f"{run.workflow_name} 差异明细"
    workbook.properties.subject = f"工作流运行 {run.id}"
    workbook.properties.creator = "Productions testing"
    output = BytesIO()
    workbook.save(output)
    timestamp = (run.finished_at or run.created_at).strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r'[\\/:*?"<>|]+', "_", run.workflow_name).strip(" ._") or "工作流"
    return output.getvalue(), f"{safe_name}_{timestamp}_差异明细.xlsx"


def _difference_row(item: WorkflowBomDifference) -> list[Any]:
    return [
        STATUS_LABELS.get(item.status, item.status),
        _excel_text(item.part_number),
        _excel_text(item.name),
        item.sop_quantity,
        item.duro_quantity,
        item.quantity_delta,
        item.sop_occurrence_count,
        _excel_text("\n".join(item.sop_locations)),
        _excel_text("\n".join(item.sop_quantity_explanations)),
        _excel_text(_occurrence_steps_text(item)),
        _excel_text("\n".join(item.duro_submenu_labels)),
        _excel_text("\n".join(item.duro_paths)),
    ]


def _ignored_row(item: WorkflowBomIgnoredItem) -> list[Any]:
    return [
        *_difference_row(item),
        IGNORE_TYPE_LABELS.get(item.ignore_type, item.ignore_type),
        _excel_text(item.ignore_value),
        _excel_text(item.ignore_reason),
        item.ignored_at.isoformat() if item.ignored_at else "",
        _excel_text(item.normalized_part_number or ""),
    ]


def _occurrence_steps_text(item: WorkflowBomDifference) -> str:
    blocks: list[str] = []
    for step in item.sop_occurrence_steps:
        delta = f"{step.quantity_delta:g}"
        if step.quantity_delta >= 0:
            delta = f"+{delta}"
        heading = f"{step.source} · 第 {step.page_number} 页 · {delta}"
        details = [heading]
        if step.action:
            details.append(f"动作：{step.action}")
        if step.evidence:
            details.append(f"正文：{step.evidence}")
        if step.reason:
            details.append(f"说明：{step.reason}")
        blocks.append("\n".join(details))
    return "\n\n".join(blocks)


def _excel_text(value: Any) -> str:
    text = str(value or "")
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _write_sheet(sheet, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="257F6B")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2E6")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = [14, 16, 24, 12, 12, 18, 18, 34, 52, 72, 24, 52, 18, 20, 34, 24, 18]
    for index, width in enumerate(widths[: len(headers)], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, sheet.max_row)}"
    sheet.row_dimensions[1].height = 24
    sheet.sheet_view.showGridLines = False
