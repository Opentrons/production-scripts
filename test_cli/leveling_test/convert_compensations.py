#!/usr/bin/env python3
"""Generate ``leveling_config.json`` from the leveling compensation workbook.

The workbook stores a measurement and an additive compensation for every
sensor channel.  For a valid row, the following must hold::

    measurement[channel] + compensation[channel] == common_reference

This module deliberately reads the channel numbers from the worksheet
headers.  The CH96 fixture does not use one global front/rear pair: C1/A2 Y
use channels 1/0 while C3 Y uses channels 3/2.  Inferring channels only from
the words ``left`` and ``right`` silently swaps C3 readings, so the converter
keeps the workbook's channel order and validates the corrected values before
writing the JSON configuration.
"""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import json
import math
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


ChannelNameMapper = Callable[[str, int, str], tuple[str, str]]
CHANNEL_HEADER_RE = re.compile(r"^(.+?)\s*\(\s*ch\s*(\d+)\s*\)\s*$", re.IGNORECASE)
FORMULA_RE = re.compile(
    r"^=\s*\$?([A-Z]{1,3})\$?(\d+)\s*([+-])\s*\$?([A-Z]{1,3})\$?(\d+)\s*$",
    re.IGNORECASE,
)
CORRECTION_TOLERANCE = 0.002


@dataclass(frozen=True)
class CompensationEntry:
    """One worksheet row mapped to one config position."""

    config_section: str
    slot_location: str
    slot_key: str
    mount: str | None
    definition_field: str
    definitions: tuple[str, ...]
    channels: tuple[int, ...]
    measurements: dict[str, float]
    compensation: dict[str, float]
    source: str

    @property
    def config_path(self) -> str:
        prefix = f"{self.config_section}.{self.slot_location}"
        return f"{prefix}.{self.mount}.{self.slot_key}" if self.mount else f"{prefix}.{self.slot_key}"


def load_json_config(filepath: str | Path) -> dict[str, Any]:
    with Path(filepath).open("r", encoding="utf-8") as file:
        return json.load(file)


def save_json_config(filepath: str | Path, data: dict[str, Any]) -> None:
    with Path(filepath).open("w", encoding="utf-8") as file:
        json.dump(data, file, indent=4, ensure_ascii=False)
        file.write("\n")


def backup_file(src_path: str | Path) -> str:
    source = Path(src_path)
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = source.with_name(f"{source.name}.backup_{timestamp}")
    shutil.copy2(source, backup_path)
    print(f"Created backup: {backup_path}")
    return str(backup_path)


def _normalise_label(value: Any) -> str:
    return re.sub(r"\s+", "_", str(value).strip().lower())


def _parse_channel_header(value: Any, source: str) -> tuple[str, int]:
    if value is None:
        raise ValueError(f"{source}: channel header is empty")
    match = CHANNEL_HEADER_RE.match(str(value).strip())
    if match is None:
        raise ValueError(f"{source}: expected '<name>(chN)' header, got {value!r}")
    return _normalise_label(match.group(1)), int(match.group(2))


def _column_letter(column: int) -> str:
    result = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _formula_value(
    formula: Any,
    values_sheet: Worksheet,
    source: str,
) -> float | None:
    """Evaluate the simple ``=cell-cell`` formulas used by old templates."""

    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    match = FORMULA_RE.match(formula.replace("$", ""))
    if match is None:
        raise ValueError(f"{source}: formula has no cached value and is unsupported: {formula!r}")
    left_ref = f"{match.group(1).upper()}{match.group(2)}"
    right_ref = f"{match.group(4).upper()}{match.group(5)}"
    left = values_sheet[left_ref].value
    right = values_sheet[right_ref].value
    if not isinstance(left, (int, float)) or not isinstance(right, (int, float)):
        raise ValueError(f"{source}: formula references non-numeric cells: {formula!r}")
    value = float(left) + (float(right) if match.group(3) == "+" else -float(right))
    return value


def _read_number(
    values_sheet: Worksheet,
    formulas_sheet: Worksheet,
    row: int,
    column: int,
    source: str,
) -> float:
    value = values_sheet.cell(row, column).value
    if isinstance(value, bool):
        raise ValueError(f"{source}: boolean is not a numeric compensation")
    if isinstance(value, (int, float)):
        formula = formulas_sheet.cell(row, column).value
        if isinstance(formula, str) and formula.startswith("="):
            evaluated = _formula_value(formula, values_sheet, source)
            if evaluated is None or not math.isclose(float(value), evaluated, abs_tol=1e-6):
                raise ValueError(
                    f"{source}: cached value {value!r} disagrees with formula {formula!r} ({evaluated!r})"
                )
        return round(float(value), 6)
    formula_value = _formula_value(formulas_sheet.cell(row, column).value, values_sheet, source)
    if formula_value is not None:
        return round(formula_value, 6)
    raise ValueError(f"{source}: expected a numeric value, got {value!r}")


def _validate_corrected_values(
    measurements: dict[str, float],
    compensation: dict[str, float],
    source: str,
) -> None:
    corrected = [measurements[name] + compensation[name] for name in measurements]
    spread = max(corrected) - min(corrected)
    if spread > CORRECTION_TOLERANCE:
        details = ", ".join(f"{name}={value:.6f}" for name, value in zip(measurements, corrected))
        raise ValueError(
            f"{source}: compensation does not equalize channels; "
            f"spread={spread:.6f} > {CORRECTION_TOLERANCE}: {details}"
        )


def _build_entry(
    values_sheet: Worksheet,
    formulas_sheet: Worksheet,
    *,
    row: int,
    header_row: int,
    measurement_columns: list[int],
    compensation_columns: list[int],
    header_columns: list[int],
    config_section: str,
    slot_location: str,
    slot_key: str,
    mount: str | None,
    definition_field: str,
    name_for_channel: ChannelNameMapper,
) -> CompensationEntry:
    if not (
        len(measurement_columns)
        == len(compensation_columns)
        == len(header_columns)
    ):
        raise ValueError(f"{values_sheet.title}: inconsistent column mapping for row {row}")

    measurements: dict[str, float] = {}
    compensation: dict[str, float] = {}
    channels: list[int] = []
    definitions: list[str] = []

    for measurement_column, compensation_column, header_column in zip(
        measurement_columns, compensation_columns, header_columns
    ):
        source = f"{values_sheet.title}!{_column_letter(header_column)}{header_row} / row {row}"
        label, channel = _parse_channel_header(values_sheet.cell(header_row, header_column).value, source)
        definition, compensation_name = name_for_channel(label, channel, source)
        if definition in definitions or compensation_name in measurements:
            raise ValueError(f"{source}: duplicate logical channel {definition!r}/{compensation_name!r}")
        definitions.append(definition)
        channels.append(channel)
        measurements[compensation_name] = _read_number(
            values_sheet,
            formulas_sheet,
            row,
            measurement_column,
            f"{values_sheet.title}!{_column_letter(measurement_column)}{row}",
        )
        compensation[compensation_name] = _read_number(
            values_sheet,
            formulas_sheet,
            row,
            compensation_column,
            f"{values_sheet.title}!{_column_letter(compensation_column)}{row}",
        )

    _validate_corrected_values(measurements, compensation, f"{values_sheet.title} row {row} ({slot_key})")
    return CompensationEntry(
        config_section=config_section,
        slot_location=slot_location,
        slot_key=slot_key,
        mount=mount,
        definition_field=definition_field,
        definitions=tuple(definitions),
        channels=tuple(channels),
        measurements=measurements,
        compensation=compensation,
        source=f"{values_sheet.title}!{row}",
    )


def _zstage_channel_name(label: str, channel: int, source: str) -> tuple[str, str]:
    expected = {
        "rear": (3, "below_rear", "rear"),
        "front": (2, "below_front", "front"),
    }
    if label not in expected or expected[label][0] != channel:
        raise ValueError(f"{source}: unexpected Z Stage channel {label}(ch{channel})")
    return expected[label][1], expected[label][2]


def _ch8_channel_name(label: str, channel: int, mount: str, source: str) -> tuple[str, str]:
    if channel not in (0, 1) or label not in ("front", "rear"):
        raise ValueError(f"{source}: unexpected 8CH channel {label}(ch{channel})")
    expected_channel = {"front": 0, "rear": 1}[label]
    if channel != expected_channel:
        raise ValueError(f"{source}: 8CH {label} must be ch{expected_channel}, got ch{channel}")
    name = f"{mount}_{label}"
    return name, name


CH96_CHANNEL_NAMES = {
    0: "left_front",
    1: "left_rear",
    2: "right_front",
    3: "right_rear",
    4: "rear_left",
    5: "rear_right",
    8: "below_front_left",
    9: "below_front_right",
    10: "below_rear_left",
    11: "below_rear_right",
}


def _ch96_y_channel_name(label: str, channel: int, source: str) -> tuple[str, str]:
    if channel not in (0, 1, 2, 3) or label not in ("front", "rear"):
        raise ValueError(f"{source}: unexpected CH96 Y channel {label}(ch{channel})")
    expected_channel = {0: "front", 1: "rear", 2: "front", 3: "rear"}[channel]
    if label != expected_channel:
        raise ValueError(f"{source}: CH96 Y channel ch{channel} is {expected_channel}, got {label}")
    name = CH96_CHANNEL_NAMES[channel]
    return name, name


def _ch96_x_channel_name(label: str, channel: int, source: str) -> tuple[str, str]:
    expected = {"left": (4, "rear_left"), "right": (5, "rear_right")}
    if label not in expected or expected[label][0] != channel:
        raise ValueError(f"{source}: unexpected CH96 X channel {label}(ch{channel})")
    name = expected[label][1]
    return name, name


def _ch96_z_channel_name(label: str, channel: int, source: str) -> tuple[str, str]:
    expected = {
        "rear_left": (10, "below_rear_left"),
        "rear_right": (11, "below_rear_right"),
        "front_left": (8, "below_front_left"),
        "front_right": (9, "below_front_right"),
    }
    if label not in expected or expected[label][0] != channel:
        raise ValueError(f"{source}: unexpected CH96 Z channel {label}(ch{channel})")
    name = expected[label][1]
    return name, name


def _read_zstage_entries(values_sheet: Worksheet, formulas_sheet: Worksheet) -> list[CompensationEntry]:
    entries: list[CompensationEntry] = []
    mount: str | None = None
    for row in range(4, values_sheet.max_row + 1):
        mount_value = _normalise_label(values_sheet.cell(row, 2).value) if values_sheet.cell(row, 2).value else ""
        if mount_value in {"left", "right"}:
            mount = mount_value
        slot_key = values_sheet.cell(row, 3).value
        if not isinstance(slot_key, str) or not slot_key.upper().startswith("Z-"):
            continue
        if mount is None:
            raise ValueError(f"Z Stage row {row}: missing Left/Right fixture section")
        entries.append(
            _build_entry(
                values_sheet,
                formulas_sheet,
                row=row,
                header_row=4 if mount == "left" else 6,
                measurement_columns=[4, 5],
                compensation_columns=[6, 7],
                header_columns=[4, 5],
                config_section="zstage_leveling_config",
                slot_location="ZStagePoint",
                slot_key=slot_key.upper(),
                mount=mount,
                definition_field="channel_definition",
                name_for_channel=_zstage_channel_name,
            )
        )
    return entries


def _read_ch8_entries(values_sheet: Worksheet, formulas_sheet: Worksheet) -> list[CompensationEntry]:
    entries: list[CompensationEntry] = []
    mount: str | None = None
    for row in range(4, values_sheet.max_row + 1):
        mount_value = _normalise_label(values_sheet.cell(row, 2).value) if values_sheet.cell(row, 2).value else ""
        if mount_value in {"left", "right"}:
            mount = mount_value
        slot_label = values_sheet.cell(row, 3).value
        if not isinstance(slot_label, str) or not slot_label.upper().startswith("Y-"):
            continue
        if mount is None:
            raise ValueError(f"8CH Pipette row {row}: missing Left/Right fixture section")
        slot_key = f"{slot_label.upper()}"
        entries.append(
            _build_entry(
                values_sheet,
                formulas_sheet,
                row=row,
                header_row=4 if mount == "left" else 6,
                measurement_columns=[4, 5],
                compensation_columns=[6, 7],
                header_columns=[6, 7],
                config_section="pipette_leveling_config",
                slot_location="SlotLocationCH8",
                slot_key=f"{slot_key}-{mount.capitalize()}",
                mount=None,
                definition_field="definition",
                name_for_channel=lambda label, channel, source, current_mount=mount: _ch8_channel_name(
                    label, channel, current_mount, source
                ),
            )
        )
    return entries


def _read_ch96_entries(values_sheet: Worksheet, formulas_sheet: Worksheet) -> list[CompensationEntry]:
    entries: list[CompensationEntry] = []
    layout: list[tuple[int, int, str, list[int], list[int], list[int], ChannelNameMapper]] = [
        (5, 4, "C1-Y", [4, 5], [8, 9], [4, 5], _ch96_y_channel_name),
        (7, 6, "C3-Y", [4, 5], [8, 9], [4, 5], _ch96_y_channel_name),
        (9, 8, "A2-Y", [4, 5], [8, 9], [4, 5], _ch96_y_channel_name),
        (11, 10, "C1-X", [4, 5], [8, 9], [4, 5], _ch96_x_channel_name),
        (12, 10, "C3-X", [4, 5], [8, 9], [4, 5], _ch96_x_channel_name),
        (13, 10, "A2-X", [4, 5], [8, 9], [4, 5], _ch96_x_channel_name),
        (15, 14, "D1-Z", [4, 5, 6, 7], [8, 9, 10, 11], [4, 5, 6, 7], _ch96_z_channel_name),
        (16, 14, "D3-Z", [4, 5, 6, 7], [8, 9, 10, 11], [4, 5, 6, 7], _ch96_z_channel_name),
        (17, 14, "C2-Z", [4, 5, 6, 7], [8, 9, 10, 11], [4, 5, 6, 7], _ch96_z_channel_name),
        (18, 14, "A2-Z", [4, 5, 6, 7], [8, 9, 10, 11], [4, 5, 6, 7], _ch96_z_channel_name),
    ]
    for row, header_row, expected_slot, measurements, compensation, headers, mapper in layout:
        actual_slot = values_sheet.cell(row, 3).value
        if _normalise_label(actual_slot) != _normalise_label(expected_slot):
            raise ValueError(
                f"96CH Pipette row {row}: expected slot {expected_slot!r}, got {actual_slot!r}"
            )
        entries.append(
            _build_entry(
                values_sheet,
                formulas_sheet,
                row=row,
                header_row=header_row,
                measurement_columns=measurements,
                compensation_columns=compensation,
                header_columns=headers,
                config_section="pipette_leveling_config",
                slot_location="SlotLocationCH96",
                slot_key=expected_slot,
                mount=None,
                definition_field="definition",
                name_for_channel=mapper,
            )
        )
    return entries


def read_excel_compensations(excel_path: str | Path) -> list[CompensationEntry]:
    """Read and validate all 27 compensation rows from the workbook."""

    path = Path(excel_path)
    values_workbook = load_workbook(path, data_only=True, read_only=False)
    formulas_workbook = load_workbook(path, data_only=False, read_only=False)
    required_sheets = {"Z Stage", "8CH Pipette", "96CH Pipette"}
    missing = required_sheets - set(values_workbook.sheetnames)
    if missing:
        raise ValueError(f"Workbook is missing required sheets: {', '.join(sorted(missing))}")

    entries = [
        *_read_zstage_entries(values_workbook["Z Stage"], formulas_workbook["Z Stage"]),
        *_read_ch8_entries(values_workbook["8CH Pipette"], formulas_workbook["8CH Pipette"]),
        *_read_ch96_entries(values_workbook["96CH Pipette"], formulas_workbook["96CH Pipette"]),
    ]
    if len(entries) != 27:
        raise ValueError(f"Expected 27 compensation rows, parsed {len(entries)}")
    return entries


def _get_position(config: dict[str, Any], entry: CompensationEntry) -> dict[str, Any]:
    section = config[entry.config_section][entry.slot_location]
    if entry.mount is not None:
        section = section[entry.mount]
    position = section.get(entry.slot_key)
    if not isinstance(position, dict):
        raise ValueError(f"Missing config position: {entry.config_path}")
    if "Point" not in position and "point" not in position:
        raise ValueError(f"Config position has no Point: {entry.config_path}")
    return position


def _channel_table(config: dict[str, Any], entry: CompensationEntry) -> dict[str, Any]:
    if entry.config_section == "zstage_leveling_config":
        mount = entry.mount
        assert mount is not None
        return config[entry.config_section]["ZStageChannel"][mount]
    mount = "left"
    if entry.slot_location == "SlotLocationCH8":
        mount = entry.slot_key.rsplit("-", 1)[-1].lower()
    return config[entry.config_section]["ChannelDefinitionCH96" if entry.slot_location == "SlotLocationCH96" else "ChannelDefinitionCH8"][mount]


def find_and_update_compensation(
    config: dict[str, Any], entries: Iterable[CompensationEntry]
) -> dict[str, dict[str, Any]]:
    """Apply validated workbook entries and return compensation/definition changes."""

    changes: dict[str, dict[str, Any]] = {}
    seen_paths: set[str] = set()
    for entry in entries:
        position = _get_position(config, entry)
        table = _channel_table(config, entry)
        for definition, channel in zip(entry.definitions, entry.channels):
            channel_config = table.get(definition)
            if not isinstance(channel_config, dict) or channel_config.get("channel") != channel:
                actual = channel_config.get("channel") if isinstance(channel_config, dict) else None
                raise ValueError(
                    f"{entry.config_path}: {definition} expected channel {channel}, got {actual}"
                )

        definition_before = position.get(entry.definition_field)
        compensation_before = position.get("compensation")
        if definition_before != list(entry.definitions) or compensation_before != entry.compensation:
            changes[entry.config_path] = {
                "definition_before": copy.deepcopy(definition_before),
                "definition_after": list(entry.definitions),
                "compensation_before": copy.deepcopy(compensation_before),
                "compensation_after": copy.deepcopy(entry.compensation),
                "source": entry.source,
            }
        position[entry.definition_field] = list(entry.definitions)
        position["compensation"] = dict(entry.compensation)
        seen_paths.add(entry.config_path)

    if len(seen_paths) != 27:
        raise ValueError(f"Expected 27 unique config positions, updated {len(seen_paths)}")
    return changes


def print_changes(changes: dict[str, dict[str, Any]]) -> None:
    if not changes:
        print("No compensation or channel-definition changes found.")
        return
    print(f"Validated {len(changes)} changed compensation/config positions:")
    for path, change in changes.items():
        print(f"  {path}")
        if change["definition_before"] != change["definition_after"]:
            print(f"    definition: {change['definition_before']} -> {change['definition_after']}")
        if change["compensation_before"] != change["compensation_after"]:
            print(f"    compensation: {change['compensation_before']} -> {change['compensation_after']}")


def _parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, default=script_dir / "Templete.xlsx")
    parser.add_argument("--config", type=Path, default=script_dir / "leveling_config.json")
    parser.add_argument("--no-backup", action="store_true", help="Do not create a config backup")
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    if not args.template.exists():
        raise SystemExit(f"Compensation template not found: {args.template}")
    if not args.config.exists():
        raise SystemExit(f"Leveling config not found: {args.config}")

    print(f"Reading compensation workbook: {args.template}")
    entries = read_excel_compensations(args.template)
    config = load_json_config(args.config)
    changes = find_and_update_compensation(config, entries)

    if not args.no_backup:
        backup_file(args.config)
    save_json_config(args.config, config)
    print(f"Generated leveling config: {args.config}")
    print_changes(changes)


if __name__ == "__main__":
    main()
