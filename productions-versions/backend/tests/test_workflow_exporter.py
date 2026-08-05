from io import BytesIO

from openpyxl import load_workbook

from workflows.exporter import build_workflow_run_workbook
from workflows.models import (
    WorkflowBomDifference,
    WorkflowBomIgnoredItem,
    WorkflowBomReport,
    WorkflowRun,
    WorkflowSopOccurrenceStep,
)


def test_workflow_run_export_contains_differences_and_ignored_sheets() -> None:
    difference = WorkflowBomDifference(
        status="quantity_mismatch",
        part_number="242-00052",
        name="扎带/zip-tie",
        sop_quantity=10,
        duro_quantity=11,
        quantity_delta=1,
        sop_occurrence_count=2,
        sop_locations=["Gen3 96ch / Assembly：第 49, 51 页"],
        sop_quantity_explanations=["料号与名称证据合计 10 根"],
        sop_occurrence_steps=[
            WorkflowSopOccurrenceStep(
                source="Gen3 96ch / Assembly",
                page_number=49,
                evidence="使用2根扎带固定",
                quantity_delta=2,
                accumulate=True,
                action="固定",
                reason="新增2根",
            )
        ],
        duro_submenu_labels=["992-00018"],
        duro_paths=["999-00196 > 992-00018 > 242-00052"],
    )
    ignored = WorkflowBomIgnoredItem(
        **difference.model_dump(),
        ignore_type="part_number",
        ignore_value="242-00052",
        ignore_reason="SOP 尚未补充完整料号",
    )
    run = WorkflowRun(
        workflow_id="workflow-test",
        workflow_name="96ch P1000 / Test",
        trigger_type="manual",
        status="succeeded",
        report=WorkflowBomReport(
            differences=[difference],
            ignored_items=[ignored],
            total_difference_count=1,
            total_ignored_count=1,
        ),
    )

    content, filename = build_workflow_run_workbook(run)
    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames == ["差异明细", "已忽略"]
    assert filename.endswith("_差异明细.xlsx")
    assert "/" not in filename
    difference_sheet = workbook["差异明细"]
    assert difference_sheet["A2"].value == "数量差异"
    assert difference_sheet["B2"].value == "242-00052"
    assert difference_sheet["J2"].value.startswith("Gen3 96ch / Assembly · 第 49 页 · +2")
    ignored_sheet = workbook["已忽略"]
    assert ignored_sheet["M2"].value == "料号"
    assert ignored_sheet["O2"].value == "SOP 尚未补充完整料号"
